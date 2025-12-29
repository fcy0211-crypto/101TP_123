import asyncio
import sqlite3
from datetime import date

from aiogram import Bot, Dispatcher, F
from aiogram.types import (
    Message, CallbackQuery,
    InlineKeyboardMarkup, InlineKeyboardButton,
    FSInputFile
)
from aiogram.filters import CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

# ================= НАСТРОЙКИ =================
BOT_TOKEN = "8397597216:AAFtzivDMoNxcRU06vp8wobfG6NU28BkIgs"
DB_NAME = "attendance.db"

STUDENTS = [
    "Бабук Владислав",
    "Гарцуев Ростислав",
    "Глинская Милена",
    "Демьянко Надежда",
    "Касьянюк Глеб",
    "Мигутский Тимур",
    "Михальчик Илья",
    "Полторако Артём",
    "Русецкая Кристина",
    "Серяков Игорь",
    "Шаболтас Матвей"
]

REASONS = [
    "по заявлению",
    "по болезни",
    "по неуважительной причине"
]

# ================= FSM =================
class AttendanceFSM(StatesGroup):
    choosing_students = State()
    choosing_reason = State()

# ================= БАЗА =================
def db():
    return sqlite3.connect(DB_NAME)

def init_db():
    with db() as conn:
        c = conn.cursor()
        c.execute("""
        CREATE TABLE IF NOT EXISTS students (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            full_name TEXT UNIQUE
        )
        """)
        c.execute("""
        CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT,
            student_id INTEGER,
            status TEXT,
            reason TEXT
        )
        """)
        for s in STUDENTS:
            c.execute(
                "INSERT OR IGNORE INTO students (full_name) VALUES (?)",
                (s,)
            )
        conn.commit()

def clear_attendance():
    with db() as conn:
        conn.execute("DELETE FROM attendance")
        conn.commit()

# ================= КЛАВИАТУРЫ =================
def main_menu():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📋 Отметить отсутствующих", callback_data="mark")],
        [InlineKeyboardButton(text="📤 Выгрузить рапортичку", callback_data="export")],
        [InlineKeyboardButton(text="🗑 Очистить рапортичку", callback_data="clear")]
    ])

def confirm_clear_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_clear"),
            InlineKeyboardButton(text="✅ Очистить", callback_data="confirm_clear")
        ]
    ])

def students_kb():
    kb = [[InlineKeyboardButton(text=s, callback_data=s)] for s in STUDENTS]
    kb.append([InlineKeyboardButton(text="✅ Готово", callback_data="done")])
    return InlineKeyboardMarkup(inline_keyboard=kb)

def reasons_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=r, callback_data=r)] for r in REASONS
    ])

# ================= ЭКСПОРТ =================
def export_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Рапортичка"

    headers = ["Дата", "ФИО", "Статус", "Причина"]
    ws.append(headers)

    for col in range(1, 5):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    with db() as conn:
        c = conn.cursor()
        c.execute("SELECT DISTINCT date FROM attendance ORDER BY date")
        dates = [d[0] for d in c.fetchall()]

        c.execute("SELECT id, full_name FROM students")
        students = c.fetchall()

        for d in dates:
            for sid, name in students:
                c.execute("""
                SELECT status, reason
                FROM attendance
                WHERE date = ? AND student_id = ?
                """, (d, sid))
                row = c.fetchone()

                if row:
                    status, reason = row
                else:
                    status, reason = "присутствовал", ""

                ws.append([d, name, status, reason])

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    table = Table(
        displayName="Attendance",
        ref=f"A1:D{ws.max_row}"
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showRowStripes=True
    )
    ws.add_table(table)

    file = "rapport_101tp.xlsx"
    wb.save(file)
    return file

# ================= BOT =================
bot = Bot(BOT_TOKEN)
dp = Dispatcher()

@dp.message(CommandStart())
async def start(message: Message):
    await message.answer(
        "📘 Рапортичка группы 101 тп",
        reply_markup=main_menu()
    )

@dp.callback_query(F.data == "mark")
async def mark(call: CallbackQuery, state: FSMContext):
    await state.update_data(date=str(date.today()))
    await state.set_state(AttendanceFSM.choosing_students)
    await call.message.answer(
        "Выберите отсутствующих:",
        reply_markup=students_kb()
    )

@dp.callback_query(AttendanceFSM.choosing_students)
async def choose_student(call: CallbackQuery, state: FSMContext):
    if call.data == "done":
        await state.clear()
        await call.message.answer("Готово ✅", reply_markup=main_menu())
        return

    await state.update_data(current_student=call.data)
    await state.set_state(AttendanceFSM.choosing_reason)

    await call.message.answer(
        f"Причина отсутствия:\n<b>{call.data}</b>",
        reply_markup=reasons_kb(),
        parse_mode="HTML"
    )

@dp.callback_query(AttendanceFSM.choosing_reason)
async def choose_reason(call: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    student = data["current_student"]
    today = data["date"]

    with db() as conn:
        c = conn.cursor()
        c.execute("SELECT id FROM students WHERE full_name = ?", (student,))
        sid = c.fetchone()[0]

        c.execute("""
        INSERT INTO attendance (date, student_id, status, reason)
        VALUES (?, ?, ?, ?)
        """, (today, sid, "отсутствовал", call.data))
        conn.commit()

    await state.set_state(AttendanceFSM.choosing_students)
    await call.message.answer(
        f"❌ {student} — {call.data}",
        reply_markup=students_kb()
    )

@dp.callback_query(F.data == "export")
async def export(call: CallbackQuery):
    file = export_excel()
    await call.message.answer_document(
        FSInputFile(file),
        caption="📤 Рапортичка группы 101 тп"
    )

# ====== ОЧИСТКА ======
@dp.callback_query(F.data == "clear")
async def clear(call: CallbackQuery):
    await call.message.answer(
        "⚠️ Вы уверены, что хотите ПОЛНОСТЬЮ очистить рапортичку?",
        reply_markup=confirm_clear_kb()
    )

@dp.callback_query(F.data == "confirm_clear")
async def confirm_clear(call: CallbackQuery):
    clear_attendance()
    await call.message.answer(
        "🗑 Рапортичка полностью очищена",
        reply_markup=main_menu()
    )

@dp.callback_query(F.data == "cancel_clear")
async def cancel_clear(call: CallbackQuery):
    await call.message.answer(
        "❌ Очистка отменена",
        reply_markup=main_menu()
    )

# ================= ЗАПУСК =================
async def main():
    init_db()
    while True:
        try:
            print("🤖 Бот запущен")
            await dp.start_polling(bot)
        except Exception as e:
            print(f"❌ Ошибка: {e}")
            print("🔄 Перезапуск через 5 секунд...")
            await asyncio.sleep(5)

if __name__ == "__main__":
    asyncio.run(main())
